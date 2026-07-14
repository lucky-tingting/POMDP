from __future__ import annotations

import argparse
import csv
import json
import math
import os
import time
from dataclasses import asdict, dataclass, replace
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from step_01_config_and_protocol import BAAMQMixRUMAConfig, GlobalTrainingState
from step_02_belief_reward import (
    expected_information_gain_torch,
    hard_observation_update_np,
    ml_soft_update_np,
    predict_belief_np,
    reward_components_from_assignment_torch,
    reward_from_assignment_torch,
)
from step_05_action_mask_and_matching import ActionMaskBuilder, CandidateSetBuilder
from step_06_trainer import BAAMQMIXRUMATrainerTorch
from step_11_training_flow import BAAMQMixRUMATrainingLoop
from step_12_online_execution_flow import online_execution_report


Point = Tuple[float, float]
ROOT = Path(__file__).resolve().parent
PACKAGE_ROOT = ROOT.parent
DEFAULT_OUTPUT_DIR = PACKAGE_ROOT / "results" / "algorithm_03_baam_qmix_ruma_torch"


@dataclass(frozen=True)
class BAAMBenchmarkExperimentConfig:
    n_points: int = 20
    n_uavs: int = 3
    horizon: int = 16
    seeds: Tuple[int, ...] = (2026, 2027, 2028, 2029, 2030)
    train_episodes: int = 24
    top_k_candidates: int = 8
    max_flight_distance: float = 22.0
    energy_capacity: float = 160.0
    energy_per_distance: float = 1.0
    min_safe_energy: float = 16.0
    grid_spacing: float = 5.0
    neighbor_radius: float = 7.2
    h_max: float = 45.0
    rainfall_base: float = 8.0
    rainfall_peak: float = 18.0
    rainfall_wave: float = 3.0
    n_ml_levels: int = 5
    ml_observation_matrix: Tuple[Tuple[float, ...], Tuple[float, ...]] = (
        (0.48, 0.25, 0.14, 0.08, 0.05),
        (0.05, 0.08, 0.14, 0.25, 0.48),
    )
    uav_sensitivity: float = 0.88
    uav_specificity: float = 0.86
    xi0: float = 0.0
    xi1: float = 3.0
    xi2: float = 1.2
    theta_01: float = 2.45
    zeta0: float = 0.0
    zeta1: float = 1.25
    zeta2: float = 1.15
    zeta3: float = 1.1
    theta_10: float = 1.15
    lambda_cover: float = 18.0
    lambda_miss: float = 8.0
    lambda_fp: float = 3.5
    lambda_cost: float = 0.18
    lambda_info: float = 3.0
    miss_threshold: float = 0.25
    gamma: float = 0.95
    learning_rate: float = 3e-4
    local_obs_dim: int = 14
    global_state_dim: int = 12
    hidden_dim: int = 48
    mixer_hidden_dim: int = 32
    target_update_interval: int = 4
    sequence_length: int = 4
    epsilon_start: float = 0.5
    epsilon_end: float = 0.05
    updates_per_episode: int = 1
    include_static_risk_features: bool = False
    include_rainfall_trend_features: bool = False
    include_freshness_feature: bool = False

    @property
    def seed_list(self) -> str:
        return ";".join(str(seed) for seed in self.seeds)

    @property
    def candidate_feature_dim(self) -> int:
        return (
            7
            + (2 if self.include_static_risk_features else 0)
            + (1 if self.include_rainfall_trend_features else 0)
            + (1 if self.include_freshness_feature else 0)
        )

    @property
    def effective_local_obs_dim(self) -> int:
        # UAV local state plus fixed Top-K candidate slots.
        return 5 + max(0, self.top_k_candidates) * self.candidate_feature_dim

    @property
    def effective_global_state_dim(self) -> int:
        # s_tr = (x, b_bar, q, E, eta), with eta represented by rainfall.
        dim = self.n_points + self.n_points + self.n_uavs * 2 + self.n_uavs + self.n_points
        if self.include_static_risk_features:
            dim += 2 * self.n_points
        if self.include_rainfall_trend_features:
            dim += self.n_points
        if self.include_freshness_feature:
            dim += self.n_points
        return dim


@dataclass(frozen=True)
class BAAMDatasetExperimentConfig(BAAMBenchmarkExperimentConfig):
    dataset_path: str = ""
    train_episodes: int = 2000
    eval_episodes: int = 2000
    eval_interval: int = 100
    hidden_dim: int = 128
    mixer_hidden_dim: int = 64
    sequence_length: int = 8
    epsilon_start: float = 1.0
    epsilon_end: float = 0.05
    updates_per_episode: int = 4
    target_update_interval: int = 100
    include_static_risk_features: bool = True
    include_rainfall_trend_features: bool = True
    include_freshness_feature: bool = True
    use_reward_curriculum: bool = True
    curriculum_core_fraction: float = 0.35
    curriculum_partial_fraction: float = 0.70
    full_experiment_profile: bool = True


def sigmoid(value: float) -> float:
    return 1.0 / (1.0 + math.exp(-value))


def distance(a: Point, b: Point) -> float:
    return math.hypot(a[0] - b[0], a[1] - b[1])


@dataclass(frozen=True)
class BAAMScenarioDataset:
    train_rainfall: np.ndarray
    train_high_risk: np.ndarray
    test_rainfall: np.ndarray
    test_high_risk: np.ndarray
    susceptibility: np.ndarray
    drainage: np.ndarray
    path: str = ""

    @classmethod
    def from_npz(cls, path: str | os.PathLike[str]) -> "BAAMScenarioDataset":
        path_obj = Path(path)
        data = np.load(path_obj)
        required = [
            "train_rainfall",
            "train_high_risk",
            "test_rainfall",
            "test_high_risk",
            "susceptibility",
            "drainage",
        ]
        missing = [key for key in required if key not in data.files]
        if missing:
            raise ValueError(f"scenario dataset is missing arrays: {missing}")
        dataset = cls(
            train_rainfall=np.asarray(data["train_rainfall"], dtype=np.float32),
            train_high_risk=np.asarray(data["train_high_risk"], dtype=np.int64),
            test_rainfall=np.asarray(data["test_rainfall"], dtype=np.float32),
            test_high_risk=np.asarray(data["test_high_risk"], dtype=np.int64),
            susceptibility=np.asarray(data["susceptibility"], dtype=np.float32),
            drainage=np.asarray(data["drainage"], dtype=np.float32),
            path=str(path_obj),
        )
        dataset.validate()
        return dataset

    def validate(self) -> None:
        for rainfall_name, state_name in [("train_rainfall", "train_high_risk"), ("test_rainfall", "test_high_risk")]:
            rainfall = getattr(self, rainfall_name)
            states = getattr(self, state_name)
            if rainfall.ndim != 3 or states.ndim != 3:
                raise ValueError(f"{rainfall_name} and {state_name} must have shape [scenario, time, point]")
            if rainfall.shape != states.shape:
                raise ValueError(f"{rainfall_name} shape {rainfall.shape} != {state_name} shape {states.shape}")
        n_points = self.train_rainfall.shape[2]
        if self.test_rainfall.shape[2] != n_points:
            raise ValueError("train/test n_points must match")
        if self.susceptibility.shape != (n_points,) or self.drainage.shape != (n_points,):
            raise ValueError("susceptibility and drainage must have shape [n_points]")

    @property
    def n_points(self) -> int:
        return int(self.train_rainfall.shape[2])

    @property
    def horizon(self) -> int:
        return int(self.train_rainfall.shape[1])

    def split_size(self, split: str) -> int:
        return int(self._rainfall(split).shape[0])

    def _rainfall(self, split: str) -> np.ndarray:
        if split == "train":
            return self.train_rainfall
        if split == "test":
            return self.test_rainfall
        raise ValueError("split must be 'train' or 'test'")

    def _states(self, split: str) -> np.ndarray:
        if split == "train":
            return self.train_high_risk
        if split == "test":
            return self.test_high_risk
        raise ValueError("split must be 'train' or 'test'")

    def scenario(self, split: str, scenario_id: int) -> tuple[np.ndarray, np.ndarray]:
        rainfall = self._rainfall(split)
        states = self._states(split)
        if scenario_id < 0 or scenario_id >= rainfall.shape[0]:
            raise IndexError(f"scenario_id {scenario_id} out of range for {split}")
        return rainfall[scenario_id].copy(), states[scenario_id].copy()


class BAAMBenchmarkEnv:
    def __init__(self, cfg: BAAMBenchmarkExperimentConfig, seed: int):
        self.cfg = cfg
        self.seed = seed
        self.reset_count = 0
        self.rng = np.random.default_rng(seed)
        self.time = 0
        self.previous_actions = np.zeros(cfg.n_uavs, dtype=np.int64)
        self.reward_weight_override: Optional[Dict[str, float]] = None
        self._build_static_space()
        self.reset()

    def _build_static_space(self) -> None:
        side = math.ceil(math.sqrt(self.cfg.n_points))
        self.locations = np.asarray(
            [
                ((i % side) * self.cfg.grid_spacing, (i // side) * self.cfg.grid_spacing)
                for i in range(self.cfg.n_points)
            ],
            dtype=np.float32,
        )
        self.names = [
            "Yuegezhuang", "Liuliqiao", "Fengyiqiao", "Dahongmen",
            "Xizhimen", "Dongzhimen", "Fuxingmen", "Jianguomen",
            "Chaoyangmen", "Fuchengmen", "Guangqumen", "Guanganmen",
            "Youanmen", "Zuoanmen", "Yongdingmen", "Andingmen",
            "Madian", "Sanyuanqiao", "Wukesong", "Shuangjing",
        ]
        self.neighbors: List[List[int]] = []
        for i, loc_i in enumerate(self.locations):
            current = []
            for j, loc_j in enumerate(self.locations):
                if i != j and distance(tuple(loc_i), tuple(loc_j)) <= self.cfg.neighbor_radius:
                    current.append(j)
            self.neighbors.append(current)

    def reset(self):
        self.rng = np.random.default_rng(self.seed + self.reset_count * 1009)
        self.reset_count += 1
        self.time = 0
        self.previous_actions = np.zeros(self.cfg.n_uavs, dtype=np.int64)
        self.uav_pos = np.zeros((self.cfg.n_uavs, 2), dtype=np.float32)
        self.uav_energy = np.full(self.cfg.n_uavs, self.cfg.energy_capacity, dtype=np.float32)
        self.susceptibility = self.rng.uniform(0.70, 1.45, size=self.cfg.n_points).astype(np.float32)
        self.drainage = self.rng.uniform(0.25, 0.95, size=self.cfg.n_points).astype(np.float32)
        self.importance = self.rng.uniform(0.80, 1.65, size=self.cfg.n_points).astype(np.float32)
        init_prob = np.minimum(0.55, 0.10 + 0.18 * self.susceptibility + 0.08 * (1.0 - self.drainage))
        self.true_state = (self.rng.random(self.cfg.n_points) < init_prob).astype(np.int64)
        low_belief = self.rng.uniform(0.18, 0.42, size=self.cfg.n_points)
        high_belief = self.rng.uniform(0.55, 0.75, size=self.cfg.n_points)
        self.belief = np.where(self.true_state == 1, high_belief, low_belief).astype(np.float32)
        self.initial_state = self.true_state.copy()
        self.initial_belief = self.belief.copy()
        self.current_rainfall = np.zeros(self.cfg.n_points, dtype=np.float32)
        self.previous_rainfall = np.zeros(self.cfg.n_points, dtype=np.float32)
        self.last_visit_age = np.full(self.cfg.n_points, self.cfg.horizon, dtype=np.float32)
        self.pre_belief = self.belief.copy()
        self.ml_levels = np.zeros(self.cfg.n_points, dtype=np.int64)
        self._prepare_decision()
        return self.local_observations()

    def _mean_neighbor_belief(self, i: int, beliefs: np.ndarray) -> float:
        if not self.neighbors[i]:
            return float(np.mean(beliefs))
        return float(np.mean(beliefs[self.neighbors[i]]))

    def _mean_neighbor_state(self, i: int) -> float:
        if not self.neighbors[i]:
            return float(np.mean(self.true_state))
        return float(np.mean(self.true_state[self.neighbors[i]]))

    def _transition_probabilities(self, i: int, rainfall: float, neighbor_mean: float) -> Tuple[float, float]:
        h_norm = min(1.5, max(0.0, rainfall * float(self.susceptibility[i]) / self.cfg.h_max))
        p01 = sigmoid(self.cfg.xi0 + self.cfg.xi1 * h_norm + self.cfg.xi2 * neighbor_mean - self.cfg.theta_01)
        p10 = sigmoid(
            self.cfg.zeta0
            + self.cfg.zeta1 * (1.0 - min(1.0, h_norm))
            + self.cfg.zeta2 * float(self.drainage[i])
            - self.cfg.zeta3 * neighbor_mean
            - self.cfg.theta_10
        )
        return min(0.72, max(0.01, p01)), min(0.65, max(0.01, p10))

    def _rainfall(self) -> np.ndarray:
        storm = self.cfg.rainfall_base + self.cfg.rainfall_peak * math.exp(
            -((self.time - self.cfg.horizon * 0.45) ** 2) / max(1.0, 0.08 * self.cfg.horizon ** 2)
        )
        wave = self.cfg.rainfall_wave * (1.0 + math.sin(2.0 * math.pi * self.time / max(1, self.cfg.horizon)))
        base = storm + wave
        rainfall = base * self.rng.uniform(0.82, 1.18, size=self.cfg.n_points) + self.rng.normal(0.0, 1.0, size=self.cfg.n_points)
        return np.maximum(0.0, rainfall).astype(np.float32)

    def _update_true_states(self) -> None:
        next_state = []
        for i in range(self.cfg.n_points):
            p01, p10 = self._transition_probabilities(i, float(self.current_rainfall[i]), self._mean_neighbor_state(i))
            if self.true_state[i] == 0:
                next_state.append(int(self.rng.random() < p01))
            else:
                next_state.append(int(not (self.rng.random() < p10)))
        self.true_state = np.asarray(next_state, dtype=np.int64)

    def _prepare_decision(self) -> None:
        self.previous_rainfall = self.current_rainfall.copy()
        self.current_rainfall = self._rainfall()
        self._update_true_states()
        p01 = np.zeros(self.cfg.n_points, dtype=np.float32)
        p10 = np.zeros(self.cfg.n_points, dtype=np.float32)
        for i in range(self.cfg.n_points):
            p01[i], p10[i] = self._transition_probabilities(
                i, float(self.current_rainfall[i]), self._mean_neighbor_belief(i, self.belief)
            )
        predicted = predict_belief_np(self.belief, p01, p10).astype(np.float32)
        likelihood = np.asarray(self.cfg.ml_observation_matrix, dtype=np.float32)
        levels = []
        for state in self.true_state:
            probs = likelihood[int(state)] / likelihood[int(state)].sum()
            levels.append(int(self.rng.choice(self.cfg.n_ml_levels, p=probs)))
        self.ml_levels = np.asarray(levels, dtype=np.int64)
        self.pre_belief = ml_soft_update_np(predicted, self.ml_levels, likelihood).astype(np.float32)

    def _space_scale(self) -> float:
        return max(1.0, self.cfg.grid_spacing * math.ceil(math.sqrt(self.cfg.n_points)))

    def _candidate_scores(self) -> np.ndarray:
        clipped = np.clip(self.pre_belief, 1e-6, 1.0 - 1e-6)
        entropy = -(clipped * np.log(clipped) + (1.0 - clipped) * np.log(1.0 - clipped))
        return CandidateSetBuilder().base_scores(
            pre_action_belief=self.pre_belief,
            point_importance=self.importance,
            entropy_value=entropy,
        )

    def _candidate_indices(self) -> np.ndarray:
        scores = self._candidate_scores()
        candidate_mask = CandidateSetBuilder().build(
            pre_action_belief=self.pre_belief,
            point_importance=self.importance,
            entropy_value=None,
            top_k=self.cfg.top_k_candidates,
        )
        ordered = [int(i) for i in np.argsort(-scores) if candidate_mask[i]]
        return np.asarray(ordered[: max(0, self.cfg.top_k_candidates)], dtype=np.int64)

    def _candidate_mask(self) -> np.ndarray:
        mask = np.zeros(self.cfg.n_points, dtype=bool)
        mask[self._candidate_indices()] = True
        return mask

    def _distance_matrix(self) -> np.ndarray:
        diff = self.uav_pos[:, None, :] - self.locations[None, :, :]
        return np.linalg.norm(diff, axis=-1).astype(np.float32)

    def _return_energy_safe_matrix(self, distance_matrix: np.ndarray) -> np.ndarray:
        depot = np.zeros(2, dtype=np.float32)
        return_distance = np.linalg.norm(self.locations - depot[None, :], axis=-1).astype(np.float32)
        total_distance = distance_matrix + return_distance[None, :]
        remaining_after_out_and_return = self.uav_energy[:, None] - self.cfg.energy_per_distance * total_distance
        return remaining_after_out_and_return >= self.cfg.min_safe_energy

    def _rainfall_trend(self) -> np.ndarray:
        return (self.current_rainfall - self.previous_rainfall).astype(np.float32)

    def _freshness_vector(self) -> np.ndarray:
        return np.clip(self.last_visit_age / max(1.0, float(self.cfg.horizon)), 0.0, 1.0).astype(np.float32)

    def local_observations(self) -> np.ndarray:
        action_masks = self.action_masks()
        feasible = action_masks[:, 1:]
        distances = self._distance_matrix()
        candidate_indices = self._candidate_indices()
        scale = self._space_scale()
        rows = []
        for m in range(self.cfg.n_uavs):
            row = [
                float(self.uav_pos[m, 0] / scale),
                float(self.uav_pos[m, 1] / scale),
                float(self.uav_energy[m] / self.cfg.energy_capacity),
                float(self.previous_actions[m] / max(1, self.cfg.n_points)),
                float(self.time / max(1, self.cfg.horizon)),
            ]
            for slot in range(max(0, self.cfg.top_k_candidates)):
                if slot < len(candidate_indices):
                    idx = int(candidate_indices[slot])
                    row.extend(
                        [
                            float(idx / max(1, self.cfg.n_points - 1)),
                            float(self.locations[idx, 0] / scale),
                            float(self.locations[idx, 1] / scale),
                            float(self.pre_belief[idx]),
                            float(self.importance[idx] / 2.0),
                            float(distances[m, idx] / max(1.0, self.cfg.max_flight_distance)),
                            float(feasible[m, idx]),
                        ]
                    )
                    if self.cfg.include_static_risk_features:
                        row.extend(
                            [
                                float(self.susceptibility[idx] / 1.5),
                                float(self.drainage[idx]),
                            ]
                        )
                    if self.cfg.include_rainfall_trend_features:
                        row.append(float(self._rainfall_trend()[idx] / max(1.0, self.cfg.h_max)))
                    if self.cfg.include_freshness_feature:
                        row.append(float(self._freshness_vector()[idx]))
                else:
                    row.extend([0.0] * self.cfg.candidate_feature_dim)
            rows.append(row)
        return np.asarray(rows, dtype=np.float32)

    def global_training_state(self) -> GlobalTrainingState:
        return GlobalTrainingState(
            true_risk_state=self.true_state.copy(),
            pre_action_belief=self.pre_belief.copy(),
            uav_positions=self.uav_pos.copy(),
            uav_energy=self.uav_energy.copy(),
            exogenous_state=self.current_rainfall.copy(),
        )

    def global_state(self) -> np.ndarray:
        base = np.concatenate(
            [
                self.true_state.astype(np.float32),
                self.pre_belief.astype(np.float32),
                (self.uav_pos / self._space_scale()).reshape(-1).astype(np.float32),
                (self.uav_energy / self.cfg.energy_capacity).astype(np.float32),
                (self.current_rainfall / self.cfg.h_max).astype(np.float32),
            ]
        ).astype(np.float32)
        extra = []
        if self.cfg.include_static_risk_features:
            extra.extend(
                [
                    (self.susceptibility / 1.5).astype(np.float32),
                    self.drainage.astype(np.float32),
                ]
            )
        if self.cfg.include_rainfall_trend_features:
            extra.append((self._rainfall_trend() / self.cfg.h_max).astype(np.float32))
        if self.cfg.include_freshness_feature:
            extra.append(self._freshness_vector())
        if extra:
            return np.concatenate([base, *extra]).astype(np.float32)
        return base

    def action_masks(self) -> np.ndarray:
        distances = self._distance_matrix()
        return ActionMaskBuilder(self._torch_cfg()).build_feasibility_mask(
            distance_matrix=distances,
            uav_energy=self.uav_energy,
            flight_time_matrix=distances,
            remaining_time=np.full(
                self.cfg.n_uavs,
                max(1, self.cfg.horizon - self.time) * self.cfg.max_flight_distance,
                dtype=np.float32,
            ),
            safe_matrix=self._return_energy_safe_matrix(distances),
            candidate_mask=self._candidate_mask(),
            max_flight_distance=np.full(self.cfg.n_uavs, self.cfg.max_flight_distance, dtype=np.float32),
            energy_per_distance=self.cfg.energy_per_distance,
            min_safe_energy=self.cfg.min_safe_energy,
        )

    def pre_action_belief(self) -> np.ndarray:
        return self.pre_belief.copy()

    def point_importance(self) -> np.ndarray:
        return self.importance.copy()

    def uav_positions(self) -> np.ndarray:
        return self.uav_pos.copy()

    def point_locations(self) -> np.ndarray:
        return self.locations.copy()

    def belief_state(self) -> np.ndarray:
        return self.belief.copy()

    def _torch_cfg(self) -> BAAMQMixRUMAConfig:
        torch_cfg = make_torch_config(self.cfg)
        if self.reward_weight_override:
            return replace(torch_cfg, **self.reward_weight_override)
        return torch_cfg

    def set_reward_weights(
        self,
        *,
        lambda_miss: Optional[float] = None,
        lambda_fp: Optional[float] = None,
        lambda_info: Optional[float] = None,
    ) -> None:
        override = {}
        if lambda_miss is not None:
            override["lambda_miss"] = float(lambda_miss)
        if lambda_fp is not None:
            override["lambda_fp"] = float(lambda_fp)
        if lambda_info is not None:
            override["lambda_info"] = float(lambda_info)
        self.reward_weight_override = override or None

    def _reward_components(self, assignment: np.ndarray, distances: np.ndarray) -> Dict[str, float]:
        components = reward_components_from_assignment_torch(
            self._torch_cfg(),
            torch.as_tensor(self.pre_belief, dtype=torch.float32),
            torch.as_tensor(self.importance, dtype=torch.float32),
            torch.as_tensor(assignment, dtype=torch.float32),
            torch.as_tensor(distances, dtype=torch.float32),
        )
        return {key: round(float(value.detach().cpu().item()), 6) for key, value in components.items()}

    def _behavior_diagnostics(self, assignment: np.ndarray, action_masks: np.ndarray) -> Dict[str, object]:
        selected = [int(i) for i in np.flatnonzero(assignment.sum(axis=0) > 0)]
        ranks = np.empty(self.cfg.n_points, dtype=np.int64)
        ranks[np.argsort(-self.pre_belief)] = np.arange(1, self.cfg.n_points + 1)
        feasible_real = action_masks[:, 1:].astype(bool)
        selected_ranks = [int(ranks[i]) for i in selected]
        high_belief = self.pre_belief >= max(self.cfg.miss_threshold, 0.5)
        masked_high = high_belief & ~feasible_real.any(axis=0)
        idle_count = int(np.sum(assignment.sum(axis=1) == 0))
        return {
            "idle_count": idle_count,
            "idle_rate": round(float(idle_count / max(1, self.cfg.n_uavs)), 6),
            "feasible_action_count_by_uav": [int(x) for x in feasible_real.sum(axis=1)],
            "feasible_action_count_mean": round(float(np.mean(feasible_real.sum(axis=1))), 6),
            "selected_belief_rank_mean": round(float(np.mean(selected_ranks)) if selected_ranks else 0.0, 6),
            "selected_mean_belief": round(float(np.mean(self.pre_belief[selected])) if selected else 0.0, 6),
            "masked_high_belief_count": int(masked_high.sum()),
        }

    def step(self, assignment_matrix: np.ndarray):
        assignment = np.asarray(assignment_matrix, dtype=np.int64)
        distances = self._distance_matrix()
        action_masks_before = self.action_masks()
        selected = [int(i) for i in np.flatnonzero(assignment.sum(axis=0) > 0)]
        hard_obs = np.full(self.cfg.n_points, -1, dtype=np.int64)
        for idx in selected:
            if self.true_state[idx] == 1:
                hard_obs[idx] = int(self.rng.random() < self.cfg.uav_sensitivity)
            else:
                hard_obs[idx] = int(self.rng.random() < (1.0 - self.cfg.uav_specificity))
        posterior = hard_observation_update_np(self._torch_cfg(), self.pre_belief, assignment, hard_obs).astype(np.float32)
        reward_components = self._reward_components(assignment, distances)
        reward = float(reward_components["total"])
        behavior_diagnostics = self._behavior_diagnostics(assignment, action_masks_before)
        high_risk = [int(i) for i in np.flatnonzero(self.true_state == 1)]
        tp = int(sum(1 for i in selected if self.true_state[i] == 1))
        fp = int(sum(1 for i in selected if self.true_state[i] == 0))
        fn = int(sum(1 for i in high_risk if i not in selected))
        route_rows = []
        for m in range(self.cfg.n_uavs):
            chosen = np.flatnonzero(assignment[m] > 0)
            start = self.uav_pos[m].copy()
            point_idx: Optional[int] = None
            dist = 0.0
            if len(chosen) > 0:
                point_idx = int(chosen[0])
                dist = float(distances[m, point_idx])
                self.uav_pos[m] = self.locations[point_idx]
                self.uav_energy[m] = max(0.0, float(self.uav_energy[m] - self.cfg.energy_per_distance * dist))
                self.previous_actions[m] = point_idx + 1
            else:
                self.previous_actions[m] = 0
            route_rows.append(
                {
                    "time": self.time,
                    "uav": m,
                    "action_index": int(self.previous_actions[m]),
                    "point_idx": point_idx,
                    "point_name": self.names[point_idx] if point_idx is not None and point_idx < len(self.names) else "",
                    "start_x": round(float(start[0]), 6),
                    "start_y": round(float(start[1]), 6),
                    "end_x": round(float(self.uav_pos[m, 0]), 6),
                    "end_y": round(float(self.uav_pos[m, 1]), 6),
                    "distance": round(dist, 6),
                    "energy_remaining": round(float(self.uav_energy[m]), 6),
                }
            )
        self.last_visit_age += 1.0
        if selected:
            self.last_visit_age[selected] = 0.0
        self.belief = posterior
        info = {
            "time": self.time,
            "rainfall": [round(float(x), 6) for x in self.current_rainfall],
            "mean_rainfall": round(float(np.mean(self.current_rainfall)), 6),
            "high_risk_set": high_risk,
            "selected_points": selected,
            "selected_names": [self.names[i] if i < len(self.names) else f"Point-{i+1}" for i in selected],
            "tp": tp,
            "fp": fp,
            "fn": fn,
            "reward": round(reward, 6),
            "reward_components": reward_components,
            "behavior_diagnostics": behavior_diagnostics,
            "pre_action_belief": [round(float(x), 6) for x in self.pre_belief],
            "posterior_belief": [round(float(x), 6) for x in posterior],
            "route_rows": route_rows,
        }
        self.time += 1
        done = self.time >= self.cfg.horizon
        if not done:
            self._prepare_decision()
        return self.local_observations(), reward, done, info

    def risk_point_rows(self, seed: int) -> List[Dict]:
        rows = []
        for i in range(self.cfg.n_points):
            rows.append(
                {
                    "seed": seed,
                    "idx": i,
                    "name": self.names[i] if i < len(self.names) else f"Point-{i+1}",
                    "x": round(float(self.locations[i, 0]), 6),
                    "y": round(float(self.locations[i, 1]), 6),
                    "flood_susceptibility": round(float(self.susceptibility[i]), 6),
                    "drainage_capacity": round(float(self.drainage[i]), 6),
                    "importance": round(float(self.importance[i]), 6),
                    "initial_state": int(self.initial_state[i]),
                    "initial_belief": round(float(self.initial_belief[i]), 6),
                    "final_state": int(self.true_state[i]),
                    "final_belief": round(float(self.belief[i]), 6),
                }
            )
        return rows


class BAAMScenarioDatasetEnv(BAAMBenchmarkEnv):
    """BAAM environment backed by scenario_dataset_v0.npz train/test splits."""

    def __init__(
        self,
        cfg: BAAMDatasetExperimentConfig,
        dataset: BAAMScenarioDataset,
        split: str,
        scenario_ids: Optional[Sequence[int]] = None,
        seed: int = 0,
    ):
        if split not in {"train", "test"}:
            raise ValueError("split must be 'train' or 'test'")
        if cfg.n_points != dataset.n_points:
            raise ValueError(f"cfg.n_points={cfg.n_points} does not match dataset n_points={dataset.n_points}")
        if cfg.horizon != dataset.horizon:
            raise ValueError(f"cfg.horizon={cfg.horizon} does not match dataset horizon={dataset.horizon}")
        self.dataset = dataset
        self.active_split = split
        self.scenario_ids = list(scenario_ids) if scenario_ids is not None else list(range(dataset.split_size(split)))
        if not self.scenario_ids:
            raise ValueError("scenario_ids must not be empty")
        self.active_scenario_id = int(self.scenario_ids[0])
        self.scenario_rainfall = np.zeros((cfg.horizon, cfg.n_points), dtype=np.float32)
        self.scenario_states = np.zeros((cfg.horizon, cfg.n_points), dtype=np.int64)
        super().__init__(cfg, seed=seed)

    def reset(self):
        self.rng = np.random.default_rng(self.seed + self.reset_count * 1009)
        scenario_pos = self.reset_count % len(self.scenario_ids)
        self.active_scenario_id = int(self.scenario_ids[scenario_pos])
        self.scenario_rainfall, self.scenario_states = self.dataset.scenario(self.active_split, self.active_scenario_id)
        self.reset_count += 1
        self.time = 0
        self.previous_actions = np.zeros(self.cfg.n_uavs, dtype=np.int64)
        self.uav_pos = np.zeros((self.cfg.n_uavs, 2), dtype=np.float32)
        self.uav_energy = np.full(self.cfg.n_uavs, self.cfg.energy_capacity, dtype=np.float32)
        self.susceptibility = self.dataset.susceptibility.copy()
        self.drainage = self.dataset.drainage.copy()
        scaled_susceptibility = (self.susceptibility - np.min(self.susceptibility)) / max(
            1e-6, float(np.ptp(self.susceptibility))
        )
        self.importance = (0.95 + 0.55 * scaled_susceptibility).astype(np.float32)
        self.current_rainfall = np.zeros(self.cfg.n_points, dtype=np.float32)
        self.previous_rainfall = np.zeros(self.cfg.n_points, dtype=np.float32)
        self.last_visit_age = np.full(self.cfg.n_points, self.cfg.horizon, dtype=np.float32)
        self.true_state = self.scenario_states[0].astype(np.int64).copy()
        low_belief = 0.18 + 0.18 * (self.current_rainfall / max(1.0, self.cfg.h_max))
        high_belief = 0.62 + 0.10 * scaled_susceptibility
        self.belief = np.where(self.true_state == 1, high_belief, low_belief).astype(np.float32)
        self.initial_state = self.true_state.copy()
        self.initial_belief = self.belief.copy()
        self.pre_belief = self.belief.copy()
        self.ml_levels = np.zeros(self.cfg.n_points, dtype=np.int64)
        self._prepare_decision()
        return self.local_observations()

    def _rainfall(self) -> np.ndarray:
        idx = min(self.time, self.scenario_rainfall.shape[0] - 1)
        return self.scenario_rainfall[idx].astype(np.float32).copy()

    def _update_true_states(self) -> None:
        idx = min(self.time, self.scenario_states.shape[0] - 1)
        self.true_state = self.scenario_states[idx].astype(np.int64).copy()

    def risk_point_rows(self, seed: int) -> List[Dict]:
        rows = super().risk_point_rows(seed)
        for row in rows:
            row["split"] = self.active_split
            row["scenario_id"] = self.active_scenario_id
        return rows


def make_torch_config(cfg: BAAMBenchmarkExperimentConfig) -> BAAMQMixRUMAConfig:
    return BAAMQMixRUMAConfig(
        n_points=cfg.n_points,
        n_uavs=cfg.n_uavs,
        local_obs_dim=cfg.effective_local_obs_dim,
        global_state_dim=cfg.effective_global_state_dim,
        hidden_dim=cfg.hidden_dim,
        mixer_hidden_dim=cfg.mixer_hidden_dim,
        gamma=cfg.gamma,
        learning_rate=cfg.learning_rate,
        target_update_interval=cfg.target_update_interval,
        lambda_cover=cfg.lambda_cover,
        lambda_miss=cfg.lambda_miss,
        lambda_fp=cfg.lambda_fp,
        lambda_cost=cfg.lambda_cost,
        lambda_info=cfg.lambda_info,
        miss_threshold=cfg.miss_threshold,
        uav_sensitivity=cfg.uav_sensitivity,
        uav_specificity=cfg.uav_specificity,
        device="cuda" if torch.cuda.is_available() else "cpu",
    )


def metric_dict(tp: int, fp: int, fn: int, rewards: Sequence[float]) -> Dict[str, float]:
    recall = tp / max(tp + fn, 1)
    precision = tp / max(tp + fp, 1)
    f1 = 2.0 * precision * recall / max(precision + recall, 1e-12)
    return {
        "tp": int(tp),
        "fp": int(fp),
        "fn": int(fn),
        "recall": round(float(recall), 6),
        "precision": round(float(precision), 6),
        "f1": round(float(f1), 6),
        "total_reward": round(float(sum(rewards)), 6),
        "mean_reward": round(float(np.mean(rewards)) if rewards else 0.0, 6),
        "total_monitored": int(tp + fp),
    }


def metric_summary(seed_rows: Sequence[Dict], cfg: BAAMBenchmarkExperimentConfig) -> Dict:
    keys = ["recall", "precision", "f1", "total_reward", "mean_reward", "total_monitored"]
    row: Dict[str, object] = {
        "algorithm": "baam_qmix_ruma_torch",
        "solver_type": "pytorch_baam_qmix_ruma_single_point_matching",
        "runs": len(seed_rows),
        "seed_list": cfg.seed_list,
        "n_points": cfg.n_points,
        "n_uavs": cfg.n_uavs,
        "horizon": cfg.horizon,
        "max_points_per_uav_per_step": 1,
        "action_definition": "one risk point per UAV per step; trajectory is accumulated across time steps",
        "shared_distance_limit": cfg.max_flight_distance,
        "shared_energy_capacity": cfg.energy_capacity,
        "shared_energy_per_distance": cfg.energy_per_distance,
        "shared_min_safe_energy": cfg.min_safe_energy,
    }
    for key in keys:
        values = np.asarray([float(r[key]) for r in seed_rows], dtype=float)
        row[key] = round(float(np.mean(values)), 6)
        row[f"{key}_std"] = round(float(np.std(values, ddof=1)) if len(values) > 1 else 0.0, 6)
    return row


def _csv_row(row: Dict) -> Dict:
    return {k: json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict, tuple)) else v for k, v in row.items()}


def write_csv(path: Path, rows: Sequence[Dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(_csv_row(row))


def append_csv_rows(path: Path, rows: Sequence[Dict], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)
    if not rows:
        if not path.exists():
            path.write_text("", encoding="utf-8")
        return
    needs_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(fieldnames), extrasaction="ignore")
        if needs_header:
            writer.writeheader()
        for row in rows:
            writer.writerow(_csv_row(row))


def write_excel(path: Path, sheets: Dict[str, Sequence[Dict] | Path]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", bold=True)
    normal_font = Font(name="Arial")
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        if isinstance(rows, Path):
            if not rows.exists() or rows.stat().st_size == 0:
                ws.append(["empty"])
                continue
            with rows.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                headers = list(reader.fieldnames or [])
                if not headers:
                    ws.append(["empty"])
                    continue
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for row in reader:
                    ws.append([row.get(h) for h in headers])
                ws.freeze_panes = "A2"
                continue
        rows = list(rows)
        if not rows:
            ws.append(["empty"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([json.dumps(row.get(h), ensure_ascii=False) if isinstance(row.get(h), (list, dict, tuple)) else row.get(h) for h in headers])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = normal_font
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = min(60, max(10, max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _evaluate_seed(cfg: BAAMBenchmarkExperimentConfig, seed: int) -> Dict[str, object]:
    torch.manual_seed(seed)
    trainer = BAAMQMIXRUMATrainerTorch(make_torch_config(cfg))
    train_env = BAAMBenchmarkEnv(cfg, seed)
    loop = BAAMQMixRUMATrainingLoop(
        trainer,
        replay_capacity_episodes=max(4, cfg.train_episodes),
        sequence_length=min(cfg.sequence_length, cfg.horizon),
        seed=seed,
    )
    training_history = loop.fit_env(
        train_env,
        episodes=cfg.train_episodes,
        horizon=cfg.horizon,
        epsilon_start=cfg.epsilon_start,
        epsilon_end=cfg.epsilon_end,
        updates_per_episode=cfg.updates_per_episode,
    )

    eval_env = BAAMBenchmarkEnv(cfg, seed + 999)
    report = online_execution_report(eval_env, trainer, cfg.horizon)
    rewards: List[float] = []
    total_tp = total_fp = total_fn = 0
    path_rows: List[Dict] = []
    rainfall_rows: List[Dict] = []
    for row in report["trace"]:
        info = dict(row["info"])
        rewards.append(float(row["reward"]))
        total_tp += int(info["tp"])
        total_fp += int(info["fp"])
        total_fn += int(info["fn"])
        for path_row in info["route_rows"]:
            out = dict(path_row)
            out["seed"] = seed
            out["reward_t"] = info["reward"]
            out["selected_points_t"] = info["selected_points"]
            path_rows.append(out)
        rainfall_rows.append(
            {
                "seed": seed,
                "time": info["time"],
                "mean_rainfall": info["mean_rainfall"],
                "rainfall": info["rainfall"],
                "high_risk_set": info["high_risk_set"],
                "high_risk_count": len(info["high_risk_set"]),
                "selected_points": info["selected_points"],
                "selected_names": info["selected_names"],
                "tp": info["tp"],
                "fp": info["fp"],
                "fn": info["fn"],
                "reward": info["reward"],
            }
        )
    metrics = metric_dict(total_tp, total_fp, total_fn, rewards)
    metrics["seed"] = seed
    return {
        "metrics": metrics,
        "training_history": training_history,
        "path_rows": path_rows,
        "rainfall_rows": rainfall_rows,
        "risk_point_rows": eval_env.risk_point_rows(seed),
    }


def config_rows(cfg: BAAMBenchmarkExperimentConfig) -> List[Dict[str, object]]:
    rows = []
    for key, value in asdict(cfg).items():
        rows.append({"parameter": key, "value": ";".join(map(str, value)) if isinstance(value, tuple) else value})
    rows.append({"parameter": "max_points_per_uav_per_step", "value": 1})
    rows.append({"parameter": "action_definition", "value": "one risk point per UAV per time step"})
    return rows


def _episode_metric_from_records(records: Sequence) -> Dict[str, float]:
    total_tp = sum(int(record.info.get("tp", 0)) for record in records)
    total_fp = sum(int(record.info.get("fp", 0)) for record in records)
    total_fn = sum(int(record.info.get("fn", 0)) for record in records)
    rewards = [float(record.reward) for record in records]
    return metric_dict(total_tp, total_fp, total_fn, rewards)


def _component_sums(records: Sequence) -> Dict[str, float]:
    keys = ["cover", "miss", "false_positive", "cost", "info", "total"]
    out = {key: 0.0 for key in keys}
    for record in records:
        components = dict(record.info.get("reward_components") or {})
        for key in keys:
            out[key] += float(components.get(key, 0.0))
    return {key: round(value, 6) for key, value in out.items()}


def _diagnostic_means(records: Sequence) -> Dict[str, float]:
    keys = [
        "idle_count",
        "idle_rate",
        "feasible_action_count_mean",
        "selected_belief_rank_mean",
        "selected_mean_belief",
        "masked_high_belief_count",
    ]
    out: Dict[str, float] = {}
    for key in keys:
        values = [float(dict(record.info.get("behavior_diagnostics") or {}).get(key, 0.0)) for record in records]
        out[key] = round(float(np.mean(values)) if values else 0.0, 6)
    return out


def _reward_override_for_episode(cfg: BAAMDatasetExperimentConfig, episode: int) -> tuple[str, Optional[Dict[str, float]]]:
    if not cfg.use_reward_curriculum:
        return "full_reward", None
    progress = episode / max(1, cfg.train_episodes - 1)
    if progress < cfg.curriculum_core_fraction:
        return "core_reward", {"lambda_miss": 0.0, "lambda_fp": 0.0, "lambda_info": 0.0}
    if progress < cfg.curriculum_partial_fraction:
        span = max(1e-9, cfg.curriculum_partial_fraction - cfg.curriculum_core_fraction)
        ratio = (progress - cfg.curriculum_core_fraction) / span
        return (
            "partial_reward",
            {
                "lambda_miss": cfg.lambda_miss * ratio,
                "lambda_fp": cfg.lambda_fp * ratio,
                "lambda_info": cfg.lambda_info * ratio,
            },
        )
    return "full_reward", None


def _flatten_step_rows(records: Sequence, seed: int, phase: str, episode: int, scenario_id: int) -> tuple[List[Dict], List[Dict]]:
    component_rows: List[Dict] = []
    diagnostic_rows: List[Dict] = []
    for t, record in enumerate(records):
        components = dict(record.info.get("reward_components") or {})
        diagnostics = dict(record.info.get("behavior_diagnostics") or {})
        component_rows.append(
            {
                "phase": phase,
                "seed": seed,
                "episode": episode,
                "scenario_id": scenario_id,
                "time": t,
                **{key: components.get(key, 0.0) for key in ["cover", "miss", "false_positive", "cost", "info", "total"]},
            }
        )
        diagnostic_rows.append(
            {
                "phase": phase,
                "seed": seed,
                "episode": episode,
                "scenario_id": scenario_id,
                "time": t,
                "idle_count": diagnostics.get("idle_count", 0),
                "idle_rate": diagnostics.get("idle_rate", 0.0),
                "feasible_action_count_by_uav": diagnostics.get("feasible_action_count_by_uav", []),
                "feasible_action_count_mean": diagnostics.get("feasible_action_count_mean", 0.0),
                "selected_belief_rank_mean": diagnostics.get("selected_belief_rank_mean", 0.0),
                "selected_mean_belief": diagnostics.get("selected_mean_belief", 0.0),
                "masked_high_belief_count": diagnostics.get("masked_high_belief_count", 0),
            }
        )
    return component_rows, diagnostic_rows


def _evaluate_dataset_policy(
    trainer: BAAMQMIXRUMATrainerTorch,
    cfg: BAAMDatasetExperimentConfig,
    dataset: BAAMScenarioDataset,
    seed: int,
    eval_episodes: int,
) -> tuple[List[Dict], List[Dict], List[Dict]]:
    scenario_ids = list(range(min(eval_episodes, dataset.split_size("test"))))
    env = BAAMScenarioDatasetEnv(cfg, dataset, split="test", scenario_ids=scenario_ids, seed=seed + 991)
    metric_rows: List[Dict] = []
    component_rows: List[Dict] = []
    diagnostic_rows: List[Dict] = []
    for episode in range(eval_episodes):
        records = []
        env.reset()
        hidden = trainer.agent.initial_hidden(trainer.cfg.n_uavs, trainer.device)
        for _ in range(cfg.horizon):
            local_obs = np.asarray(env.local_observations(), dtype=np.float32)
            action = trainer.act(
                local_obs=local_obs,
                hidden_state=hidden,
                action_mask=np.asarray(env.action_masks(), dtype=bool),
                pre_action_belief=np.asarray(env.pre_action_belief(), dtype=np.float32),
                point_importance=np.asarray(env.point_importance(), dtype=np.float32),
                uav_positions=np.asarray(env.uav_positions(), dtype=np.float32),
                point_locations=np.asarray(env.point_locations(), dtype=np.float32),
            )
            hidden = action["next_hidden"]
            _next_obs, reward, done, info = env.step(np.asarray(action["assignment_matrix"], dtype=np.int64))
            records.append(type("EvalRecord", (), {"reward": reward, "info": dict(info or {})})())
            if done:
                break
        metrics = _episode_metric_from_records(records)
        metric_rows.append(
            {
                "phase": "test",
                "seed": seed,
                "episode": episode,
                "scenario_id": env.active_scenario_id,
                **metrics,
                **_component_sums(records),
                **_diagnostic_means(records),
            }
        )
        comp, diag = _flatten_step_rows(records, seed, "test", episode, env.active_scenario_id)
        component_rows.extend(comp)
        diagnostic_rows.extend(diag)
    return metric_rows, component_rows, diagnostic_rows


def _summarize_test_metrics(rows: Sequence[Dict], cfg: BAAMDatasetExperimentConfig) -> Dict[str, object]:
    summary: Dict[str, object] = {
        "algorithm": "baam_qmix_ruma_torch_dataset",
        "dataset_path": cfg.dataset_path,
        "train_episodes": cfg.train_episodes,
        "eval_episodes": cfg.eval_episodes,
        "seeds": cfg.seed_list,
        "n_points": cfg.n_points,
        "n_uavs": cfg.n_uavs,
        "horizon": cfg.horizon,
        "full_experiment_profile": cfg.full_experiment_profile,
        "use_reward_curriculum": cfg.use_reward_curriculum,
    }
    for key in ["recall", "precision", "f1", "total_reward", "mean_reward", "total_monitored"]:
        values = np.asarray([float(row[key]) for row in rows], dtype=float)
        summary[key] = round(float(np.mean(values)) if len(values) else 0.0, 6)
        summary[f"{key}_std"] = round(float(np.std(values, ddof=1)) if len(values) > 1 else 0.0, 6)
    return summary


def run_baam_qmix_ruma_dataset_experiment(
    cfg: BAAMDatasetExperimentConfig,
    output_dir: Optional[str | os.PathLike[str]] = None,
) -> Dict[str, object]:
    if not cfg.dataset_path:
        raise ValueError("dataset_path is required for dataset experiment")
    dataset = BAAMScenarioDataset.from_npz(cfg.dataset_path)
    output = Path(output_dir) if output_dir is not None else DEFAULT_OUTPUT_DIR / "dataset_experiment"
    output.mkdir(parents=True, exist_ok=True)
    output_files = {
        "summary": "baam_dataset_summary.csv",
        "train_curve": "baam_dataset_train_curve.csv",
        "test_metrics": "baam_dataset_test_metrics.csv",
        "reward_components": "baam_dataset_reward_components.csv",
        "behavior_diagnostics": "baam_dataset_behavior_diagnostics.csv",
        "progress": "baam_dataset_progress.csv",
        "config": "baam_dataset_config.csv",
        "excel": "baam_dataset_results.xlsx",
        "checkpoint": "baam_dataset_checkpoint.pt",
        "json": "baam_dataset_result.json",
    }
    reward_component_path = output / output_files["reward_components"]
    behavior_diagnostic_path = output / output_files["behavior_diagnostics"]
    progress_path = output / output_files["progress"]
    for stale_path in [reward_component_path, behavior_diagnostic_path, progress_path]:
        if stale_path.exists():
            stale_path.unlink()

    train_curve_rows: List[Dict] = []
    test_metric_rows: List[Dict] = []
    progress_rows: List[Dict] = []
    trainers: List[BAAMQMIXRUMATrainerTorch] = []
    started_at = time.time()
    reward_component_fields = [
        "phase",
        "seed",
        "episode",
        "scenario_id",
        "time",
        "cover",
        "miss",
        "false_positive",
        "cost",
        "info",
        "total",
    ]
    behavior_diagnostic_fields = [
        "phase",
        "seed",
        "episode",
        "scenario_id",
        "time",
        "idle_count",
        "idle_rate",
        "feasible_action_count_by_uav",
        "feasible_action_count_mean",
        "selected_belief_rank_mean",
        "selected_mean_belief",
        "masked_high_belief_count",
    ]

    def log_progress(seed: int, phase: str, episode: int, row: Optional[Dict] = None) -> None:
        row = dict(row or {})
        progress_row = {
            "elapsed_seconds": round(time.time() - started_at, 3),
            "seed": seed,
            "phase": phase,
            "episode": int(episode),
            "train_episodes": cfg.train_episodes,
            "eval_episodes": cfg.eval_episodes,
            "seeds": cfg.seed_list,
            "reward_mode": row.get("reward_mode", ""),
            "recall": row.get("recall", ""),
            "precision": row.get("precision", ""),
            "f1": row.get("f1", ""),
            "total_reward": row.get("total_reward", ""),
            "loss": row.get("loss", ""),
        }
        progress_rows.append(progress_row)
        write_csv(progress_path, progress_rows)
        print(
            "progress "
            f"seed={seed} phase={phase} episode={episode} "
            f"elapsed={progress_row['elapsed_seconds']}s "
            f"f1={progress_row['f1']} reward={progress_row['total_reward']}",
            flush=True,
        )

    for seed in cfg.seeds:
        torch.manual_seed(seed)
        trainer = BAAMQMIXRUMATrainerTorch(make_torch_config(cfg))
        trainers.append(trainer)
        log_progress(seed, "seed_start", 0)
        train_env = BAAMScenarioDatasetEnv(
            cfg,
            dataset,
            split="train",
            scenario_ids=list(range(dataset.split_size("train"))),
            seed=seed,
        )
        loop = BAAMQMixRUMATrainingLoop(
            trainer,
            replay_capacity_episodes=max(64, min(cfg.train_episodes, 10000)),
            sequence_length=min(cfg.sequence_length, cfg.horizon),
            seed=seed,
        )
        for episode in range(cfg.train_episodes):
            reward_mode, override = _reward_override_for_episode(cfg, episode)
            if override is None:
                train_env.set_reward_weights()
            else:
                train_env.set_reward_weights(**override)
            epsilon = max(
                cfg.epsilon_end,
                cfg.epsilon_start - (cfg.epsilon_start - cfg.epsilon_end) * episode / max(1, cfg.train_episodes - 1),
            )
            records = loop.collect_episode(train_env, horizon=cfg.horizon, epsilon=epsilon)
            metrics = _episode_metric_from_records(records)
            row: Dict[str, object] = {
                "phase": "train",
                "seed": seed,
                "episode": episode,
                "scenario_id": train_env.active_scenario_id,
                "epsilon": round(float(epsilon), 6),
                "reward_mode": reward_mode,
                **metrics,
                **_component_sums(records),
                **_diagnostic_means(records),
            }
            if loop.replay.can_sample():
                update_metrics: Dict[str, float] = {}
                for _ in range(cfg.updates_per_episode):
                    update_metrics = loop.update_from_replay(batch_size=1)
                row.update(update_metrics)
            train_curve_rows.append(row)
            comp, diag = _flatten_step_rows(records, seed, "train", episode, train_env.active_scenario_id)
            append_csv_rows(reward_component_path, comp, reward_component_fields)
            append_csv_rows(behavior_diagnostic_path, diag, behavior_diagnostic_fields)
            if (episode + 1) % max(1, cfg.eval_interval) == 0 or episode == cfg.train_episodes - 1:
                log_progress(seed, "train", episode + 1, row)

        eval_rows, eval_components, eval_diagnostics = _evaluate_dataset_policy(
            trainer,
            cfg,
            dataset,
            seed,
            eval_episodes=cfg.eval_episodes,
        )
        test_metric_rows.extend(eval_rows)
        append_csv_rows(reward_component_path, eval_components, reward_component_fields)
        append_csv_rows(behavior_diagnostic_path, eval_diagnostics, behavior_diagnostic_fields)
        if eval_rows:
            log_progress(seed, "test_done", cfg.eval_episodes, eval_rows[-1])

    summary = _summarize_test_metrics(test_metric_rows, cfg)
    cfg_rows = config_rows(cfg)
    write_csv(output / output_files["summary"], [summary])
    write_csv(output / output_files["train_curve"], train_curve_rows)
    write_csv(output / output_files["test_metrics"], test_metric_rows)
    write_csv(output / output_files["progress"], progress_rows)
    write_csv(output / output_files["config"], cfg_rows)
    write_excel(
        output / output_files["excel"],
        {
            "summary": [summary],
            "train_curve": train_curve_rows,
            "test_metrics": test_metric_rows,
            "reward_components": reward_component_path,
            "behavior_diagnostics": behavior_diagnostic_path,
            "progress": progress_rows,
            "config": cfg_rows,
        },
    )
    if trainers:
        trainers[-1].save(output / output_files["checkpoint"])
    result = {
        "summary_metrics": summary,
        "output_dir": str(output),
        "output_files": output_files,
    }
    (output / output_files["json"]).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def run_baam_qmix_ruma_experiment(
    cfg: Optional[BAAMBenchmarkExperimentConfig] = None,
    output_dir: Optional[str | os.PathLike[str]] = None,
) -> Dict[str, object]:
    cfg = cfg or BAAMBenchmarkExperimentConfig()
    output = Path(output_dir) if output_dir is not None else DEFAULT_OUTPUT_DIR
    output.mkdir(parents=True, exist_ok=True)

    seed_results = [_evaluate_seed(cfg, seed) for seed in cfg.seeds]
    seed_metrics = [dict(item["metrics"]) for item in seed_results]
    summary = metric_summary(seed_metrics, cfg)
    path_rows = [row for item in seed_results for row in item["path_rows"]]
    rainfall_rows = [row for item in seed_results for row in item["rainfall_rows"]]
    risk_rows = [row for item in seed_results for row in item["risk_point_rows"]]
    cfg_rows = config_rows(cfg)

    output_files = {
        "summary": "baam_qmix_ruma_summary.csv",
        "seed_metrics": "baam_qmix_ruma_seed_metrics.csv",
        "uav_paths": "baam_qmix_ruma_uav_paths.csv",
        "rainfall_high_risk": "baam_qmix_ruma_rainfall_high_risk.csv",
        "risk_points": "baam_qmix_ruma_risk_points.csv",
        "config": "baam_qmix_ruma_config.csv",
        "json": "baam_qmix_ruma_result.json",
        "excel": "baam_qmix_ruma_results.xlsx",
    }
    write_csv(output / output_files["summary"], [summary])
    write_csv(output / output_files["seed_metrics"], seed_metrics)
    write_csv(output / output_files["uav_paths"], path_rows)
    write_csv(output / output_files["rainfall_high_risk"], rainfall_rows)
    write_csv(output / output_files["risk_points"], risk_rows)
    write_csv(output / output_files["config"], cfg_rows)
    write_excel(
        output / output_files["excel"],
        {
            "summary": [summary],
            "seed_metrics": seed_metrics,
            "uav_paths": path_rows,
            "rainfall_high_risk": rainfall_rows,
            "risk_points": risk_rows,
            "config": cfg_rows,
        },
    )
    result = {
        "summary_metrics": summary,
        "seed_metrics": seed_metrics,
        "output_dir": str(output),
        "output_files": output_files,
    }
    (output / output_files["json"]).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def _parse_seed_tuple(value: str) -> Tuple[int, ...]:
    return tuple(int(part.strip()) for part in value.split(",") if part.strip())


def main(argv: Optional[Sequence[str]] = None) -> None:
    parser = argparse.ArgumentParser(description="Run BAAM-QMIX-RUMA experiments.")
    parser.add_argument("--dataset-experiment", action="store_true", help="Run the real scenario-dataset training/testing pipeline.")
    parser.add_argument("--dataset-path", default="", help="Path to scenario_dataset_v0.npz. Defaults to results/hpr/scenario_v0/scenario_dataset_v0.npz.")
    parser.add_argument("--output-dir", default="", help="Directory for experiment outputs.")
    parser.add_argument("--train-episodes", type=int, default=None, help="Training episodes for dataset experiment.")
    parser.add_argument("--eval-episodes", type=int, default=None, help="Held-out test episodes for dataset experiment.")
    parser.add_argument("--seeds", default="", help="Comma-separated seed list, for example 2026,2027,2028.")
    parser.add_argument("--no-curriculum", action="store_true", help="Disable core-to-full reward curriculum.")
    args = parser.parse_args(argv)

    if args.dataset_experiment:
        default_dataset = PACKAGE_ROOT / "results" / "hpr" / "scenario_v0" / "scenario_dataset_v0.npz"
        dataset_path = args.dataset_path or str(default_dataset)
        cfg_kwargs: Dict[str, object] = {"dataset_path": dataset_path}
        if args.train_episodes is not None:
            cfg_kwargs["train_episodes"] = args.train_episodes
        if args.eval_episodes is not None:
            cfg_kwargs["eval_episodes"] = args.eval_episodes
        if args.seeds:
            cfg_kwargs["seeds"] = _parse_seed_tuple(args.seeds)
        if args.no_curriculum:
            cfg_kwargs["use_reward_curriculum"] = False
        cfg = BAAMDatasetExperimentConfig(**cfg_kwargs)
        output = args.output_dir or str(DEFAULT_OUTPUT_DIR / "dataset_experiment")
        result = run_baam_qmix_ruma_dataset_experiment(cfg, output)
        summary = result["summary_metrics"]
        print("BAAM-QMIX-RUMA dataset experiment completed.")
        print(
            f"train_episodes={summary['train_episodes']}, eval_episodes={summary['eval_episodes']}, "
            f"recall={summary['recall']:.3f}, precision={summary['precision']:.3f}, "
            f"f1={summary['f1']:.3f}, reward={summary['total_reward']:.2f}"
        )
        print(f"Results saved to: {result['output_dir']}")
        return

    result = run_baam_qmix_ruma_experiment(output_dir=args.output_dir or None)
    summary = result["summary_metrics"]
    print("BAAM-QMIX-RUMA PyTorch experiment completed.")
    print(
        f"runs={summary['runs']}, N={summary['n_points']}, M={summary['n_uavs']}, T={summary['horizon']}, "
        f"recall={summary['recall']:.3f}, precision={summary['precision']:.3f}, "
        f"f1={summary['f1']:.3f}, reward={summary['total_reward']:.2f}"
    )
    print(f"Results saved to: {result['output_dir']}")


if __name__ == "__main__":
    main()
