from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent


class Algorithm03ExperimentRunnerTests(unittest.TestCase):
    def test_default_config_uses_shared_energy_proxy_settings(self):
        from run_baam_qmix_ruma_experiment import BAAMBenchmarkExperimentConfig

        cfg = BAAMBenchmarkExperimentConfig()

        self.assertEqual(cfg.n_points, 20)
        self.assertEqual(cfg.n_uavs, 3)
        self.assertEqual(cfg.horizon, 16)
        self.assertEqual(cfg.seeds, (2026, 2027, 2028, 2029, 2030))
        self.assertEqual(cfg.max_flight_distance, 22.0)
        self.assertEqual(cfg.energy_capacity, 160.0)
        self.assertEqual(cfg.energy_per_distance, 1.0)
        self.assertEqual(cfg.min_safe_energy, 16.0)

    def test_runner_uses_shared_public_experiment_parameters_and_exports_excel(self):
        from run_baam_qmix_ruma_experiment import BAAMBenchmarkExperimentConfig, run_baam_qmix_ruma_experiment

        cfg = BAAMBenchmarkExperimentConfig(
            n_points=8,
            n_uavs=2,
            horizon=3,
            seeds=(31, 32),
            train_episodes=2,
            top_k_candidates=4,
            hidden_dim=16,
            mixer_hidden_dim=8,
        )
        with tempfile.TemporaryDirectory() as tmp:
            result = run_baam_qmix_ruma_experiment(cfg, tmp)
            output_dir = Path(tmp)

            self.assertEqual(result["summary_metrics"]["runs"], 2)
            self.assertEqual(result["summary_metrics"]["n_points"], 8)
            self.assertEqual(result["summary_metrics"]["n_uavs"], 2)
            self.assertEqual(result["summary_metrics"]["horizon"], 3)
            self.assertEqual(result["summary_metrics"]["max_points_per_uav_per_step"], 1)
            self.assertEqual(result["summary_metrics"]["seed_list"], "31;32")

            required = [
                "baam_qmix_ruma_summary.csv",
                "baam_qmix_ruma_seed_metrics.csv",
                "baam_qmix_ruma_uav_paths.csv",
                "baam_qmix_ruma_rainfall_high_risk.csv",
                "baam_qmix_ruma_risk_points.csv",
                "baam_qmix_ruma_result.json",
                "baam_qmix_ruma_results.xlsx",
            ]
            for name in required:
                self.assertTrue((output_dir / name).exists(), name)

            wb = load_workbook(output_dir / "baam_qmix_ruma_results.xlsx", read_only=True, data_only=False)
            try:
                self.assertEqual(
                    set(wb.sheetnames),
                    {"summary", "seed_metrics", "uav_paths", "rainfall_high_risk", "risk_points", "config"},
                )
            finally:
                wb.close()

    def test_benchmark_global_state_uses_full_training_vector(self):
        from run_baam_qmix_ruma_experiment import BAAMBenchmarkEnv, BAAMBenchmarkExperimentConfig, make_torch_config

        cfg = BAAMBenchmarkExperimentConfig(
            n_points=5,
            n_uavs=2,
            horizon=3,
            seeds=(11,),
            train_episodes=1,
            top_k_candidates=5,
        )
        env = BAAMBenchmarkEnv(cfg, seed=11)

        state = env.global_state()
        expected_dim = cfg.n_points + cfg.n_points + cfg.n_uavs * 2 + cfg.n_uavs + cfg.n_points

        self.assertEqual(state.shape, (expected_dim,))
        self.assertEqual(make_torch_config(cfg).global_state_dim, expected_dim)
        self.assertEqual(state[: cfg.n_points].tolist(), env.true_state.astype("float32").tolist())

    def test_benchmark_local_observation_contains_candidate_slots(self):
        from run_baam_qmix_ruma_experiment import BAAMBenchmarkEnv, BAAMBenchmarkExperimentConfig, make_torch_config

        cfg = BAAMBenchmarkExperimentConfig(
            n_points=6,
            n_uavs=2,
            horizon=3,
            seeds=(12,),
            train_episodes=1,
            top_k_candidates=3,
        )
        env = BAAMBenchmarkEnv(cfg, seed=12)

        obs = env.local_observations()
        expected_dim = 5 + cfg.top_k_candidates * 7

        self.assertEqual(obs.shape, (cfg.n_uavs, expected_dim))
        self.assertEqual(make_torch_config(cfg).local_obs_dim, expected_dim)

    def test_action_mask_enforces_return_energy_safety(self):
        from run_baam_qmix_ruma_experiment import BAAMBenchmarkEnv, BAAMBenchmarkExperimentConfig

        cfg = BAAMBenchmarkExperimentConfig(
            n_points=5,
            n_uavs=1,
            horizon=3,
            seeds=(13,),
            train_episodes=1,
            top_k_candidates=5,
            grid_spacing=10.0,
            max_flight_distance=100.0,
            energy_capacity=30.0,
            energy_per_distance=1.0,
            min_safe_energy=10.0,
        )
        env = BAAMBenchmarkEnv(cfg, seed=13)
        env.uav_energy[:] = 30.0

        mask = env.action_masks()

        self.assertTrue(mask[0, 0])
        self.assertFalse(mask[0, 3], "point 2 is reachable outbound but unsafe when return energy is required")


if __name__ == "__main__":
    unittest.main()
