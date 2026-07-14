from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

import numpy as np
import torch
from openpyxl import load_workbook


def make_tiny_dataset(path: Path) -> None:
    rng = np.random.default_rng(123)
    train_rainfall = rng.uniform(1.0, 20.0, size=(3, 4, 5)).astype(np.float32)
    test_rainfall = rng.uniform(1.0, 20.0, size=(2, 4, 5)).astype(np.float32)
    train_high_risk = np.zeros((3, 4, 5), dtype=np.int8)
    test_high_risk = np.zeros((2, 4, 5), dtype=np.int8)
    train_high_risk[:, :, [1, 3]] = 1
    test_high_risk[:, :, [0, 4]] = 1
    susceptibility = np.linspace(0.8, 1.4, 5, dtype=np.float32)
    drainage = np.linspace(0.3, 0.9, 5, dtype=np.float32)
    np.savez(
        path,
        train_rainfall=train_rainfall,
        train_high_risk=train_high_risk,
        test_rainfall=test_rainfall,
        test_high_risk=test_high_risk,
        susceptibility=susceptibility,
        drainage=drainage,
    )


class RealExperimentPipelineTests(unittest.TestCase):
    def test_matching_keeps_one_to_one_assignment_constraints(self):
        from step_01_config_and_protocol import BAAMQMixRUMAConfig
        from step_05_action_mask_and_matching import RollingMatcher
        from step_11_training_flow import random_feasible_matching

        cfg = BAAMQMixRUMAConfig(n_points=4, n_uavs=3)
        action_mask = np.ones((cfg.n_uavs, cfg.n_actions), dtype=bool)

        for seed in range(20):
            assignment, action_indices = random_feasible_matching(action_mask, seed=seed)
            self.assertTrue(np.all(assignment.sum(axis=1) <= 1))
            self.assertTrue(np.all(assignment.sum(axis=0) <= 1))
            self.assertEqual(action_indices.shape, (cfg.n_uavs,))

        matcher = RollingMatcher(cfg)
        q_values = torch.tensor(
            [
                [0.0, 9.0, 8.0, 7.0, 1.0],
                [0.0, 10.0, 9.0, 1.0, 1.0],
                [0.0, 11.0, 1.0, 8.0, 1.0],
            ],
            dtype=torch.float32,
        )
        assignment, _assignments, _scores = matcher.select_assignment(
            q_values=q_values,
            action_mask=torch.as_tensor(action_mask),
            pre_action_belief=torch.full((cfg.n_points,), 0.5),
            point_importance=torch.ones(cfg.n_points),
            uav_positions=torch.zeros((cfg.n_uavs, 2)),
            point_locations=torch.zeros((cfg.n_points, 2)),
        )
        self.assertTrue(np.all(assignment.sum(axis=1) <= 1))
        self.assertTrue(np.all(assignment.sum(axis=0) <= 1))

    def test_dataset_env_uses_train_and_test_splits_without_mixing(self):
        from run_baam_qmix_ruma_experiment import (
            BAAMDatasetExperimentConfig,
            BAAMScenarioDataset,
            BAAMScenarioDatasetEnv,
        )

        with tempfile.TemporaryDirectory() as tmp:
            dataset_path = Path(tmp) / "tiny_scenario_dataset.npz"
            make_tiny_dataset(dataset_path)
            dataset = BAAMScenarioDataset.from_npz(dataset_path)
            cfg = BAAMDatasetExperimentConfig(
                dataset_path=str(dataset_path),
                n_points=5,
                n_uavs=2,
                horizon=4,
                top_k_candidates=3,
                train_episodes=1,
            )

            train_env = BAAMScenarioDatasetEnv(cfg, dataset, split="train", scenario_ids=[1], seed=1)
            test_env = BAAMScenarioDatasetEnv(cfg, dataset, split="test", scenario_ids=[0], seed=1)

            train_env.reset()
            test_env.reset()

            np.testing.assert_allclose(train_env.current_rainfall, dataset.train_rainfall[1, 0])
            np.testing.assert_array_equal(train_env.true_state, dataset.train_high_risk[1, 0])
            np.testing.assert_allclose(test_env.current_rainfall, dataset.test_rainfall[0, 0])
            np.testing.assert_array_equal(test_env.true_state, dataset.test_high_risk[0, 0])
            self.assertEqual(train_env.active_split, "train")
            self.assertEqual(test_env.active_split, "test")
            self.assertEqual(train_env.global_state().shape, (cfg.effective_global_state_dim,))
            self.assertEqual(train_env.local_observations().shape, (cfg.n_uavs, cfg.effective_local_obs_dim))

    def test_step_info_contains_reward_decomposition_and_behavior_diagnostics(self):
        from run_baam_qmix_ruma_experiment import (
            BAAMDatasetExperimentConfig,
            BAAMScenarioDataset,
            BAAMScenarioDatasetEnv,
        )

        with tempfile.TemporaryDirectory() as tmp:
            dataset_path = Path(tmp) / "tiny_scenario_dataset.npz"
            make_tiny_dataset(dataset_path)
            dataset = BAAMScenarioDataset.from_npz(dataset_path)
            cfg = BAAMDatasetExperimentConfig(
                dataset_path=str(dataset_path),
                n_points=5,
                n_uavs=2,
                horizon=4,
                top_k_candidates=5,
                train_episodes=1,
                max_flight_distance=100.0,
            )
            env = BAAMScenarioDatasetEnv(cfg, dataset, split="train", scenario_ids=[0], seed=2)
            env.reset()
            assignment = np.zeros((cfg.n_uavs, cfg.n_points), dtype=np.int64)
            assignment[0, 1] = 1

            _obs, reward, _done, info = env.step(assignment)

            components = info["reward_components"]
            diagnostics = info["behavior_diagnostics"]
            self.assertEqual(
                set(components),
                {"cover", "miss", "false_positive", "cost", "info", "total"},
            )
            self.assertAlmostEqual(float(components["total"]), float(reward), places=5)
            self.assertIn("idle_count", diagnostics)
            self.assertIn("idle_rate", diagnostics)
            self.assertIn("feasible_action_count_mean", diagnostics)
            self.assertIn("selected_belief_rank_mean", diagnostics)
            self.assertIn("masked_high_belief_count", diagnostics)

    def test_dataset_experiment_exports_training_curve_test_metrics_diagnostics_and_checkpoint(self):
        from run_baam_qmix_ruma_experiment import BAAMDatasetExperimentConfig, run_baam_qmix_ruma_dataset_experiment

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            dataset_path = tmp_path / "tiny_scenario_dataset.npz"
            make_tiny_dataset(dataset_path)
            cfg = BAAMDatasetExperimentConfig(
                dataset_path=str(dataset_path),
                n_points=5,
                n_uavs=2,
                horizon=4,
                seeds=(7,),
                train_episodes=2,
                eval_episodes=2,
                top_k_candidates=3,
                hidden_dim=8,
                mixer_hidden_dim=4,
                sequence_length=2,
                updates_per_episode=1,
                eval_interval=1,
                full_experiment_profile=True,
            )

            result = run_baam_qmix_ruma_dataset_experiment(cfg, tmp_path / "outputs")
            output_dir = Path(result["output_dir"])

            expected = [
                "baam_dataset_train_curve.csv",
                "baam_dataset_test_metrics.csv",
                "baam_dataset_reward_components.csv",
                "baam_dataset_behavior_diagnostics.csv",
                "baam_dataset_progress.csv",
                "baam_dataset_results.xlsx",
                "baam_dataset_checkpoint.pt",
            ]
            for name in expected:
                self.assertTrue((output_dir / name).exists(), name)

            self.assertEqual(result["summary_metrics"]["train_episodes"], 2)
            self.assertEqual(result["summary_metrics"]["eval_episodes"], 2)
            self.assertEqual(result["summary_metrics"]["dataset_path"], str(dataset_path))
            wb = load_workbook(output_dir / "baam_dataset_results.xlsx", read_only=True, data_only=False)
            try:
                self.assertIn("train_curve", wb.sheetnames)
                self.assertIn("test_metrics", wb.sheetnames)
                self.assertIn("reward_components", wb.sheetnames)
                self.assertIn("behavior_diagnostics", wb.sheetnames)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
