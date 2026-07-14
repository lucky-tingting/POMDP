from __future__ import annotations

import csv
import json
import os
import sys
from dataclasses import asdict, dataclass
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ALG1_DIR = os.path.join(ROOT, "algorithm_01_rolling_route_optimization")
ALG2_DIR = os.path.join(ROOT, "algorithm_02_deep_qmix_gru_path_planning")
ALG3_DIR = os.path.join(ROOT, "algorithm_03_baam_qmix_ruma_torch")
ALG4_DIR = os.path.join(ROOT, "algorithm_04_new_solver_for_comparison")
RESULTS_DIR = os.path.join(ROOT, "results")

sys.path.insert(0, ALG1_DIR)
sys.path.insert(0, ALG2_DIR)
sys.path.insert(0, ALG3_DIR)
sys.path.insert(0, ALG4_DIR)

from mf_mle_muav_uwr_pomdp import ModelConfig
from run_rolling_route_exports import run_rolling_route_algorithm
from deep_qmix_gru_path_planning import PaperQMIXConfig, run_deep_qmix_multi_seed
from run_baam_qmix_ruma_experiment import BAAMBenchmarkExperimentConfig, run_baam_qmix_ruma_experiment
from step_01_config_and_parameters import PPBVIRUMAConfig
from step_06_run_p_pbvi_ruma_experiment import run_algorithm_04_experiment
from experiment_result_index import build_experiment_result_index


@dataclass(frozen=True)
class SharedExperimentSettings:
    """Shared settings that must be identical for both comparison algorithms."""

    n_points: int = 20
    n_uavs: int = 3
    horizon: int = 16
    seeds: Tuple[int, ...] = (2026, 2027, 2028, 2029, 2030)
    qmix_episodes: int = 24
    qmix_route_length: int = 1
    max_route_distance: float = 22.0
    energy_capacity: float = 160.0
    energy_per_distance: float = 1.0
    min_safe_energy: float = 16.0
    lambda_cover: float = 18.0
    lambda_miss: float = 8.0
    lambda_fp: float = 3.5
    lambda_cost: float = 0.18
    lambda_info: float = 3.0
    miss_threshold: float = 0.25

    @property
    def seed_list(self) -> str:
        return ";".join(str(seed) for seed in self.seeds)


def build_rolling_config(settings: SharedExperimentSettings) -> ModelConfig:
    return ModelConfig(
        n_points=settings.n_points,
        n_uavs=settings.n_uavs,
        horizon=settings.horizon,
        seed=settings.seeds[0],
        max_flight_distance=settings.max_route_distance,
        energy_capacity=settings.energy_capacity,
        energy_per_distance=settings.energy_per_distance,
        min_safe_energy=settings.min_safe_energy,
        lambda_cover=settings.lambda_cover,
        lambda_miss=settings.lambda_miss,
        lambda_fp=settings.lambda_fp,
        lambda_cost=settings.lambda_cost,
        lambda_info=settings.lambda_info,
        miss_threshold=settings.miss_threshold,
    )


def build_qmix_config(settings: SharedExperimentSettings) -> PaperQMIXConfig:
    return PaperQMIXConfig(
        n_points=settings.n_points,
        n_uavs=settings.n_uavs,
        horizon=settings.horizon,
        episodes=settings.qmix_episodes,
        route_length=settings.qmix_route_length,
        max_route_distance=settings.max_route_distance,
        energy_capacity=settings.energy_capacity,
        energy_per_distance=settings.energy_per_distance,
        min_safe_energy=settings.min_safe_energy,
        lambda_cover=settings.lambda_cover,
        lambda_miss=settings.lambda_miss,
        lambda_fp=settings.lambda_fp,
        lambda_cost=settings.lambda_cost,
        lambda_info=settings.lambda_info,
        miss_threshold=settings.miss_threshold,
        seed=settings.seeds[0],
    )


def build_baam_config(settings: SharedExperimentSettings) -> BAAMBenchmarkExperimentConfig:
    return BAAMBenchmarkExperimentConfig(
        n_points=settings.n_points,
        n_uavs=settings.n_uavs,
        horizon=settings.horizon,
        seeds=settings.seeds,
        train_episodes=settings.qmix_episodes,
        max_flight_distance=settings.max_route_distance,
        energy_capacity=settings.energy_capacity,
        energy_per_distance=settings.energy_per_distance,
        min_safe_energy=settings.min_safe_energy,
        lambda_cover=settings.lambda_cover,
        lambda_miss=settings.lambda_miss,
        lambda_fp=settings.lambda_fp,
        lambda_cost=settings.lambda_cost,
        lambda_info=settings.lambda_info,
        miss_threshold=settings.miss_threshold,
    )


def build_ppbvi_config(settings: SharedExperimentSettings) -> PPBVIRUMAConfig:
    return PPBVIRUMAConfig(
        n_points=settings.n_points,
        n_uavs=settings.n_uavs,
        horizon=settings.horizon,
        seeds=settings.seeds,
        max_route_distance=settings.max_route_distance,
        energy_capacity=settings.energy_capacity,
        energy_per_distance=settings.energy_per_distance,
        min_safe_energy=settings.min_safe_energy,
        lambda_cover=settings.lambda_cover,
        lambda_miss=settings.lambda_miss,
        lambda_fp=settings.lambda_fp,
        lambda_cost=settings.lambda_cost,
        lambda_info=settings.lambda_info,
        miss_threshold=settings.miss_threshold,
    )


def write_csv(path: str, rows: Sequence[Dict]) -> None:
    if not rows:
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write("")
        return
    fieldnames = list(rows[0].keys())
    for row in rows[1:]:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel(path: str, sheets: Dict[str, Sequence[Dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", bold=True)
    normal_font = Font(name="Arial")
    for sheet_name, rows in sheets.items():
        rows = list(rows)
        ws = wb.create_sheet(sheet_name[:31])
        if not rows:
            ws.append(["empty"])
            continue
        headers = list(rows[0].keys())
        for row in rows[1:]:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(header) for header in headers])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = normal_font
        ws.freeze_panes = "A2"
        for column in ws.columns:
            width = min(60, max(10, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width
    wb.save(path)


def settings_rows(settings: SharedExperimentSettings) -> List[Dict[str, object]]:
    rows = []
    for key, value in asdict(settings).items():
        rows.append({"parameter": key, "value": ";".join(map(str, value)) if isinstance(value, tuple) else value})
    rows.append({"parameter": "max_points_per_uav_per_step", "value": 1})
    rows.append({"parameter": "action_definition", "value": "one risk point per UAV per time step"})
    return rows


def write_run_summary_markdown(path: str, rows: Sequence[Dict], settings: SharedExperimentSettings) -> None:
    lines = [
        "# 四算法统一运行与对比",
        "",
        "统一入口：",
        "",
        "```powershell",
        f"python \"{os.path.join(ROOT, 'run_all.py')}\"",
        "```",
        "",
        "当前公共实验设置：",
        "",
        f"- N = {settings.n_points}",
        f"- M = {settings.n_uavs}",
        f"- T = {settings.horizon}",
        f"- seeds = {settings.seed_list}",
        f"- 每架无人机每个时间步最多访问 1 个风险点",
        f"- 距离约束 = {settings.max_route_distance}",
        f"- 能量代理约束 = capacity {settings.energy_capacity}, per_distance {settings.energy_per_distance}, min_safe {settings.min_safe_energy}",
        "",
        "统一汇总结果：",
        "",
        "- `results/comparison_summary.csv`",
        "- `results/comparison_summary.xlsx`",
        "- `results/combined_result.json`",
        "",
        "| algorithm | runs | action capacity | recall | precision | f1 | total_reward |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        lines.append(
            f"| {row['algorithm']} | {row['runs']} | {row['max_points_per_uav_per_step']} | "
            f"{row['recall']:.6f} +/- {row['recall_std']:.6f} | "
            f"{row['precision']:.6f} +/- {row['precision_std']:.6f} | "
            f"{row['f1']:.6f} +/- {row['f1_std']:.6f} | "
            f"{row['total_reward']:.6f} +/- {row['total_reward_std']:.6f} |"
        )
    lines.extend(
        [
            "",
            "说明：四算法现在统一的是 N/M/T/seeds、单步动作容量、奖励权重、距离约束和距离-能耗代理约束。",
            "algorithm_03 与 algorithm_04 都使用统一公共参数，但它们的逐时 rainfall/high-risk/belief/selected-points 历史数据由各自 runner 生成，不是同一份实时 scenario。",
            "真实能耗模型、避障、防撞、返航安全约束和空域规则仍未完整实现；当前代码对应距离/能耗代理约束下的仿真实验。",
        ]
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def _metric(row: Dict, key: str) -> float:
    return row.get(key, row.get(f"{key}_mean", 0.0))


def _metric_std(row: Dict, key: str) -> float:
    return row.get(f"{key}_std", 0.0)


def comparison_rows(
    rolling_result: Dict,
    qmix_result: Dict,
    baam_result: Dict,
    ppbvi_result: Dict,
    settings: SharedExperimentSettings,
) -> List[Dict]:
    rolling_best = next(row for row in rolling_result["summary"] if row["policy"] == "rolling_matching")
    qmix_metrics = qmix_result["summary_metrics"]
    baam_metrics = baam_result["summary_metrics"]
    ppbvi_metrics = ppbvi_result["summary_metrics"]
    return [
        {
            "algorithm": "rolling_route_optimization",
            "solver_type": "modeling_level_receding_horizon_matching",
            "comparison_role": "baseline_single-point_receding-horizon_matching",
            "runs": rolling_best["runs"],
            "seed_list": settings.seed_list,
            "n_points": settings.n_points,
            "n_uavs": settings.n_uavs,
            "horizon": settings.horizon,
            "max_points_per_uav_per_step": 1,
            "action_definition": "one risk point per UAV per step; multi-step trajectory is accumulated over time",
            "shared_distance_limit": settings.max_route_distance,
            "shared_energy_capacity": settings.energy_capacity,
            "shared_energy_per_distance": settings.energy_per_distance,
            "shared_min_safe_energy": settings.min_safe_energy,
            "recall": _metric(rolling_best, "recall"),
            "recall_std": _metric_std(rolling_best, "recall"),
            "precision": _metric(rolling_best, "precision"),
            "precision_std": _metric_std(rolling_best, "precision"),
            "f1": _metric(rolling_best, "f1"),
            "f1_std": _metric_std(rolling_best, "f1"),
            "total_reward": _metric(rolling_best, "total_reward"),
            "total_reward_std": _metric_std(rolling_best, "total_reward"),
            "mean_reward": _metric(rolling_best, "mean_reward"),
            "mean_reward_std": _metric_std(rolling_best, "mean_reward"),
            "total_monitored": _metric(rolling_best, "total_monitored"),
            "total_monitored_std": _metric_std(rolling_best, "total_monitored"),
            "fairness_note": "Shared N/M/T/seeds/reward parameters/distance and energy proxy constraints/action capacity; Deep trajectory is accumulated across time steps.",
            "path_output": "results/algorithm_01_rolling_route_optimization/rolling_modeling_uav_paths.csv",
            "rainfall_output": "results/algorithm_01_rolling_route_optimization/rolling_modeling_rainfall_high_risk.csv",
            "risk_point_output": "results/algorithm_01_rolling_route_optimization/rolling_modeling_risk_points.csv",
        },
        {
            "algorithm": "deep_qmix_gru_path_planning",
            "solver_type": "deep_marl_qmix_gru_ctde_single_point_actions",
            "comparison_role": "main_qmix_gru_single-point_matching_actions",
            "runs": rolling_best["runs"],
            "seed_list": settings.seed_list,
            "n_points": settings.n_points,
            "n_uavs": settings.n_uavs,
            "horizon": settings.horizon,
            "max_points_per_uav_per_step": settings.qmix_route_length,
            "action_definition": "one risk point per UAV per step; the UAV trajectory is accumulated across time steps",
            "shared_distance_limit": settings.max_route_distance,
            "shared_energy_capacity": settings.energy_capacity,
            "shared_energy_per_distance": settings.energy_per_distance,
            "shared_min_safe_energy": settings.min_safe_energy,
            "recall": _metric(qmix_metrics, "recall"),
            "recall_std": _metric_std(qmix_metrics, "recall"),
            "precision": _metric(qmix_metrics, "precision"),
            "precision_std": _metric_std(qmix_metrics, "precision"),
            "f1": _metric(qmix_metrics, "f1"),
            "f1_std": _metric_std(qmix_metrics, "f1"),
            "total_reward": _metric(qmix_metrics, "total_reward"),
            "total_reward_std": _metric_std(qmix_metrics, "total_reward"),
            "mean_reward": _metric(qmix_metrics, "mean_reward"),
            "mean_reward_std": _metric_std(qmix_metrics, "mean_reward"),
            "total_monitored": _metric(qmix_metrics, "total_monitored"),
            "total_monitored_std": _metric_std(qmix_metrics, "total_monitored"),
            "fairness_note": "Shared N/M/T/seeds/reward parameters/distance and energy proxy constraints/action capacity; Deep trajectory is accumulated across time steps.",
            "path_output": "results/algorithm_02_deep_qmix_gru_path_planning/deep_qmix_gru_uav_paths.csv",
            "rainfall_output": "results/algorithm_02_deep_qmix_gru_path_planning/deep_qmix_gru_rainfall_high_risk.csv",
            "risk_point_output": "results/algorithm_02_deep_qmix_gru_path_planning/deep_qmix_gru_risk_points.csv",
        },
        {
            "algorithm": "baam_qmix_ruma_torch",
            "solver_type": "pytorch_baam_qmix_ruma_single_point_matching",
            "comparison_role": "PyTorch BAAM-QMIX-RUMA single-point matching implementation",
            "runs": baam_metrics["runs"],
            "seed_list": settings.seed_list,
            "n_points": settings.n_points,
            "n_uavs": settings.n_uavs,
            "horizon": settings.horizon,
            "max_points_per_uav_per_step": 1,
            "action_definition": "one risk point per UAV per step; the UAV trajectory is accumulated across time steps",
            "shared_distance_limit": settings.max_route_distance,
            "shared_energy_capacity": settings.energy_capacity,
            "shared_energy_per_distance": settings.energy_per_distance,
            "shared_min_safe_energy": settings.min_safe_energy,
            "recall": _metric(baam_metrics, "recall"),
            "recall_std": _metric_std(baam_metrics, "recall"),
            "precision": _metric(baam_metrics, "precision"),
            "precision_std": _metric_std(baam_metrics, "precision"),
            "f1": _metric(baam_metrics, "f1"),
            "f1_std": _metric_std(baam_metrics, "f1"),
            "total_reward": _metric(baam_metrics, "total_reward"),
            "total_reward_std": _metric_std(baam_metrics, "total_reward"),
            "mean_reward": _metric(baam_metrics, "mean_reward"),
            "mean_reward_std": _metric_std(baam_metrics, "mean_reward"),
            "total_monitored": _metric(baam_metrics, "total_monitored"),
            "total_monitored_std": _metric_std(baam_metrics, "total_monitored"),
            "fairness_note": "Shared N/M/T/seeds/reward parameters/distance and energy proxy constraints/action capacity; BAAM-QMIX-RUMA uses PyTorch training and action masking.",
            "path_output": "results/algorithm_03_baam_qmix_ruma_torch/baam_qmix_ruma_uav_paths.csv",
            "rainfall_output": "results/algorithm_03_baam_qmix_ruma_torch/baam_qmix_ruma_rainfall_high_risk.csv",
            "risk_point_output": "results/algorithm_03_baam_qmix_ruma_torch/baam_qmix_ruma_risk_points.csv",
        },
        {
            "algorithm": "p_pbvi_ruma",
            "solver_type": "Pointwise-PBVI Index-based Rolling Multi-UAV Matching",
            "comparison_role": "algorithm_04_pointwise_pbvi_index_matching",
            "runs": ppbvi_metrics["runs"],
            "seed_list": settings.seed_list,
            "n_points": settings.n_points,
            "n_uavs": settings.n_uavs,
            "horizon": settings.horizon,
            "max_points_per_uav_per_step": 1,
            "action_definition": "one risk point per UAV per step; the UAV trajectory is accumulated across time steps",
            "shared_distance_limit": settings.max_route_distance,
            "shared_energy_capacity": settings.energy_capacity,
            "shared_energy_per_distance": settings.energy_per_distance,
            "shared_min_safe_energy": settings.min_safe_energy,
            "recall": _metric(ppbvi_metrics, "recall"),
            "recall_std": _metric_std(ppbvi_metrics, "recall"),
            "precision": _metric(ppbvi_metrics, "precision"),
            "precision_std": _metric_std(ppbvi_metrics, "precision"),
            "f1": _metric(ppbvi_metrics, "f1"),
            "f1_std": _metric_std(ppbvi_metrics, "f1"),
            "total_reward": _metric(ppbvi_metrics, "total_reward"),
            "total_reward_std": _metric_std(ppbvi_metrics, "total_reward"),
            "mean_reward": _metric(ppbvi_metrics, "mean_reward"),
            "mean_reward_std": _metric_std(ppbvi_metrics, "mean_reward"),
            "total_monitored": _metric(ppbvi_metrics, "total_monitored"),
            "total_monitored_std": _metric_std(ppbvi_metrics, "total_monitored"),
            "fairness_note": "Shared N/M/T/seeds/reward parameters/distance and energy proxy constraints/action capacity; P-PBVI-RUMA uses pointwise PBVI indices and rolling matching.",
            "path_output": "results/algorithm_04_new_solver_for_comparison/algorithm_04_uav_paths.csv",
            "rainfall_output": "results/algorithm_04_new_solver_for_comparison/algorithm_04_rainfall_high_risk.csv",
            "risk_point_output": "results/algorithm_04_new_solver_for_comparison/algorithm_04_risk_points.csv",
            "matching_output": "results/algorithm_04_new_solver_for_comparison/algorithm_04_matching_matrices.csv",
            "belief_output": "results/algorithm_04_new_solver_for_comparison/algorithm_04_belief_sequence.csv",
        },
    ]


def write_main_readme(path: str, rows: Sequence[Dict]) -> None:
    lines = [
        "# MF-MLe-MUAV-UWR-POMDP 三算法统一求解包",
        "",
        "本文件夹把三套求解算法分开保存，并提供统一主函数。",
        "",
        "## 三个求解算法到底是什么",
        "",
        "1. `algorithm_01_rolling_route_optimization`：论文建模级滚动优化算法。",
        "   它基于动作前信念、风险点重要性、漏检惩罚、误监测惩罚、信息价值和飞行成本构造边权，逐时间步求最大权匹配。它不是深度学习算法，优点是可解释、稳定，适合作为论文基准求解器。",
        "",
        "2. `algorithm_02_deep_qmix_gru_path_planning`：深度 QMIX-GRU + CTDE 的动态监测轨迹算法。",
        "   按最终建模版，每架无人机每个时间步最多访问 1 个风险点；跨多个时间步累计后形成无人机轨迹 `rho^m=(i_1^m,i_2^m,...,i_T^m)`。训练时使用全局状态、共享奖励和单调 mixer；执行时每架无人机只使用自己的局部观测和 GRU 隐状态选择当前时间步动作。",
        "",
        "## 怎么跑",
        "",
        "在 PowerShell 或命令行运行：",
        "",
        "```powershell",
        f"python \"{os.path.join(ROOT, 'run_all.py')}\"",
        "```",
        "",
        "也可以分别运行：",
        "",
        "```powershell",
        f"python \"{os.path.join(ALG1_DIR, 'run_rolling_route_exports.py')}\"",
        f"python \"{os.path.join(ALG2_DIR, 'deep_qmix_gru_path_planning.py')}\"",
        "```",
        "",
        "## 怎么对比好坏",
        "",
        "核心指标如下：",
        "",
        "- `Recall = TP/(TP+FN)`：高风险点发现率，越高越说明漏检少。",
        "- `Precision = TP/(TP+FP)`：监测命中率，越高越说明误监测少。",
        "- `F1`：Recall 和 Precision 的折中指标。",
        "- `total_reward`：综合覆盖收益、漏检惩罚、误监测惩罚、飞行成本和信息价值后的总目标值。",
        "- `uav_paths.csv`：检查每架无人机每个时间步到底走了哪条路径，不能只看总指标。",
        "",
        "严格论文对比必须统一可统一的实验设置：`N/M/T/seeds/奖励参数/降雨生成机制/风险点生成机制/距离-能耗代理约束`。当前总入口已经统一这些公共设置，并对三算法都使用相同 5 个随机种子报告均值和标准差。",
        "",
        "<span style=\"color:red\"><strong>最新修正后，三算法的动作容量已经统一：每架无人机每个时间步最多访问 1 个点。完整路径由多个时间步动作累计形成，这与最终建模版一致。</strong></span>",
        "",
        "## 路径、降雨、风险点参数在哪里",
        "",
        "| 算法 | 路径文件 | 降雨/高风险过程 | 固定风险点参数 |",
        "|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            f"| {row['algorithm']} | `{row['path_output']}` | `{row['rainfall_output']}` | `{row['risk_point_output']}` |"
        )
    lines.extend(
        [
            "",
            "## 本次运行汇总（多随机种子均值±标准差）",
            "",
            "| algorithm | runs | action capacity | recall | precision | f1 | total_reward | mean_reward |",
            "|---|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for row in rows:
        lines.append(
            f"| {row['algorithm']} | {row['runs']} | {row['max_points_per_uav_per_step']} | "
            f"{row['recall']:.6f}±{row['recall_std']:.6f} | "
            f"{row['precision']:.6f}±{row['precision_std']:.6f} | "
            f"{row['f1']:.6f}±{row['f1_std']:.6f} | "
            f"{row['total_reward']:.6f}±{row['total_reward_std']:.6f} | "
            f"{row['mean_reward']:.6f}±{row['mean_reward_std']:.6f} |"
        )
    lines.extend(
        [
            "",
            "## 重要说明",
            "",
            "这里说的“论文级”是指代码结构已经对应论文模型要素：固定候选点、动态高风险集合、ML 软观测、无人机硬观测、均值场信念传播、路径动作、CTDE/QMIX-GRU 或可解释滚动优化。若要做最终论文大规模实验，应继续扩大 `N`，并根据真实北京市风险点数据替换模拟风险点。",
            "",
            "## 当前约束同步状态",
            "",
            "<span style=\"color:red\"><strong>当前三算法已经同步了公共实验参数、单步动作容量、距离可行性约束和距离-能耗代理约束；但最终建模版没有显式给出避障、防撞、返航安全和空域规则的数学约束，因此这些工程约束没有被强行加入。</strong></span>",
            "",
            "最终建模版中明确出现的是飞行距离/飞行成本 `D(q_t^m, ell_i)`，以及泛化的可行动作集合 `A(q_t)` 中的续航、任务范围等约束。当前代码实现的是距离线性能耗代理约束，不是真实电池模型；若后续论文要加入避障、防撞、返航或空域规则，需要先把这些约束写成明确数学公式，再同步到三算法。",
        ]
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def write_main_readme(path: str, rows: Sequence[Dict]) -> None:
    return


def run_all(settings: SharedExperimentSettings = SharedExperimentSettings()) -> Dict:
    os.makedirs(RESULTS_DIR, exist_ok=True)
    rolling_dir = os.path.join(RESULTS_DIR, "algorithm_01_rolling_route_optimization")
    qmix_dir = os.path.join(RESULTS_DIR, "algorithm_02_deep_qmix_gru_path_planning")
    baam_dir = os.path.join(RESULTS_DIR, "algorithm_03_baam_qmix_ruma_torch")
    ppbvi_dir = os.path.join(RESULTS_DIR, "algorithm_04_new_solver_for_comparison")
    os.makedirs(rolling_dir, exist_ok=True)
    os.makedirs(qmix_dir, exist_ok=True)
    os.makedirs(baam_dir, exist_ok=True)
    os.makedirs(ppbvi_dir, exist_ok=True)

    rolling_cfg = build_rolling_config(settings)
    qmix_cfg = build_qmix_config(settings)
    baam_cfg = build_baam_config(settings)
    ppbvi_cfg = build_ppbvi_config(settings)
    rolling_result = run_rolling_route_algorithm(rolling_dir, cfg=rolling_cfg, seeds=settings.seeds)
    qmix_result = run_deep_qmix_multi_seed(qmix_cfg, qmix_dir, seeds=settings.seeds)
    baam_result = run_baam_qmix_ruma_experiment(baam_cfg, baam_dir)
    ppbvi_result = run_algorithm_04_experiment(ppbvi_cfg, ppbvi_dir)

    rows = comparison_rows(rolling_result, qmix_result, baam_result, ppbvi_result, settings)
    write_csv(os.path.join(RESULTS_DIR, "comparison_summary.csv"), rows)
    write_excel(
        os.path.join(RESULTS_DIR, "comparison_summary.xlsx"),
        {
            "comparison_summary": rows,
            "shared_settings": settings_rows(settings),
        },
    )
    write_run_summary_markdown(os.path.join(ROOT, "00_三算法统一运行与对比.md"), rows, settings)
    write_run_summary_markdown(os.path.join(ROOT, "00_四算法统一运行与对比.md"), rows, settings)
    with open(os.path.join(RESULTS_DIR, "combined_result.json"), "w", encoding="utf-8") as f:
        json.dump({"settings": asdict(settings), "comparison": rows}, f, ensure_ascii=False, indent=2)
    index_counts = build_experiment_result_index()
    return {"settings": asdict(settings), "comparison": rows, "index_counts": index_counts}


def main() -> None:
    result = run_all()
    print("MF-MLe-MUAV-UWR-POMDP four-solver package completed.")
    for row in result["comparison"]:
        print(
            f"{row['algorithm']}: recall={row['recall']:.3f}, precision={row['precision']:.3f}, "
            f"f1={row['f1']:.3f}, reward={row['total_reward']:.2f}"
        )
    print(f"Results saved to: {RESULTS_DIR}")


if __name__ == "__main__":
    main()
